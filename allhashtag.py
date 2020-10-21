# -*- coding: utf-8 -*-

from bs4 import BeautifulSoup
import selenium.webdriver as webdriver
import urllib.parse
from urllib.request import Request, urlopen
from time import sleep
import pandas as pd
from selenium.webdriver.common.keys import Keys
import time
import datetime
import traceback

import numpy as np

import openpyxl

import configparser

# reallink = []

def getList():
    reallink = []
    while True:
        pageString = driver.page_source
        bsObj = BeautifulSoup(pageString, "lxml")

        for link1 in bsObj.find_all(name="div",attrs={"class":"Nnq7C weEfm"}):
            # print("len link1:", len(link1.select('a')))
            for idx in range(len(link1.select('a'))):
                title = link1.select('a')[idx]
                real = title.attrs['href']
                reallink.append(real)


        last_height = driver.execute_script("return document.body.scrollHeight")
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        sleep(SCROLL_PAUSE_TIME)
        new_height = driver.execute_script("return document.body.scrollHeight")
        if new_height == last_height:
            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            sleep(SCROLL_PAUSE_TIME)
            new_height = driver.execute_script("return document.body.scrollHeight")
            if new_height == last_height:
                break

            else:
                last_height = new_height
                continue



    myset = set(reallink)
    reallink = list(myset)
    ilink = np.array(reallink)
    np.save("link"+name, ilink)
    print('### 저장완료 ', name)


if __name__ == "__main__":

    config = configparser.ConfigParser()
    config.read('config.ini')

    username = config['DEFAULT']['USERID']
    userpw = config['DEFAULT']['PASSWD']

    MAC = 0
    driver = ""

    wb = openpyxl.Workbook()
    sheet = wb.active
    # 진주, 거창x, 창원x

    #1 인천(168),
    #  평택(368)
    #2 부산(491),
    #3 김해(579),
    #4 안동(593),
    #5 대구(616),
    #6 제천(709),
    #7 대전(719)
    #8 울산(757),
    #9 제주(949),
    #10 군산(1137),
    #11 삼척(1146)
    #12 양산(1197)
    #13 수원(1198)
    #14 천안(1391),
    #15 금산(1945),
    #16 여수(4218)
    #17 진주(6302)
    #18 전주(10998)


    name = "천안청년몰"  #input("검색어를 입력하세요 : " )
    print('### 검색어 ', name)

    # sheet = wb.create_sheet(name)
    # csvtext = [['erna_limdaugh', 'photozone', '청년몰', 'arirangtv', '전주여행', 'handmadestores', 'rangmate', 'touristspot']]
    # value = ",".join(csvtext[0])
    # print("value: " , value)
    #

    if (MAC):
        # driver load
        driver = webdriver.Chrome(executable_path='./chromedriver')
    else:
        driver = webdriver.Chrome('chromedriver.exe')

    driver.set_window_size(600,600)

    loginUrl = 'https://www.instagram.com/accounts/login/'
    driver.implicitly_wait(5)

    # 웹 사이트 접속
    driver.get(loginUrl)

    # 사전 정보 정의


    #login
    login_id = driver.find_element_by_name('username')
    login_id.send_keys(username)
    login_pw = driver.find_element_by_name('password')
    login_pw.send_keys(userpw)
    login_pw.send_keys(Keys.RETURN)
    time.sleep(5)

    popup = driver.find_element_by_xpath('//*[@id="react-root"]/section/main/div/div/div/div/button')
    popup.send_keys(Keys.ENTER)
    time.sleep(3)
    popup = driver.find_element_by_xpath('/html/body/div[4]/div/div/div/div[3]/button[2]')
    popup.send_keys(Keys.ENTER)


    scrolltime = 1.5 #float(input("크롤링 속도를 입력하세요 : "))
    # crawlnum = int(input("가져올 데이터의 수를 입력하세요 : " ))


    # sheet = wb.create_sheet(name)


    search = urllib.parse.quote(name)
    url = 'https://www.instagram.com/explore/tags/'+str(search)+'/'
    # driver = webdriver.Chrome('chromedriver.exe')




    driver.get(url)
    sleep(5)


    now = datetime.datetime.now()
    SCROLL_PAUSE_TIME = 2.0

    reallink = []

    getList()

    reallink = np.load('link'+name+'.npy')

    # abc = np.load('link'+name+'.npy')
    # reallink = abc[1001:]

    hashtags2 = []
    totalLikes = 0

    # ilinks = ",".join(reallink)
    # print("ilinks: " , ilinks)
    # sheet.append([ilinks])
    # sheet.append([reallink])
    #

    # wb.save("link"+name + nowDatetime + ".xlsx")

    reallinknum = len(reallink)
    print("총"+str(reallinknum)+"개의 데이터.")
    sheet.append([str(reallinknum)])

    try:  # 반복문 시작 ( print 명령어로 원하는 문자열인지 하나씩 확인해보시길 바랍니다.
        for i in range(0,reallinknum):
            hashtags2.append([])
            percen = ( i / reallinknum * 100)
            print(" count: (%d  @/ %d)" % (percen, i))
            driver.implicitly_wait(1)
            req = 'https://www.instagram.com/'+reallink[i]
            driver.get(req)
            # print("Links: ", req)
            webpage = driver.page_source
            soup = BeautifulSoup(webpage, "html.parser")
            #print(soup)
            # day = str(soup.find_all(name="time",attrs={"class":"_1o9PC Nzb55"}))
            # date = day.split('일">')[1].split('</time>')[0]
            # # print("date:", date)

            soup1 = str(soup.find_all(attrs={'class': 'e1e1d'}))
            #print(soup1)
            user_id = soup1.split('href="/')[1].split('/"')[0]
            #print(user_id)
            soup1 = str(soup.find_all(attrs={'class': 'Nm9Fw'}))
            #print(soup1)
            subValue = 'span'
            if(soup1=="[]"): #좋아요가 0개, 1개, n개 일경우 모두 소스가 다르다.
                likes = '0'
            elif( soup1.find(subValue)==-1):
                likes = '0' #soup1.split('좋아요 ')[1].split('개')[0]
            elif( soup1.find(subValue)!=-1):
                likes = soup1.split('<span>')[1].split('</span>')[0]

            # totalLikes += int(likes)

            insert_data = {
                # "date" :  date,
                "likes" :  likes
            }
            # pandas_csv.to_csv(insert_data)

            soup1 = str(soup.find_all(attrs={'class': 'xil3i'}))
            if(soup1=="[]"): #해쉬태그가 없을 경우 소스가 다르다.
                hashtags = '해쉬태그없음'
                insert_data = {
                                "user_id" : user_id,
                                "좋아요" : likes}
                # pan_csv.to_csv(insert_data)
                # print("해쉬태그없음: ", insert_data)
            else:
                soup2 = soup1.split(',')
                soup2num = len(soup2)
                for j in range(0,soup2num):
                    hashtags = soup2[j].split('#')[1].split('</a>')[0]
                    # print("hashtags: ", hashtags)
                    insert_data = { "hashtags" : hashtags }
                    # print("저장성공: %s", insert_data)
                    sheet.append([hashtags])

                # soup2 = soup1.split(',')
                # soup2num = len(soup2)
                # csvtext = []
                # # csvtext.append(str(i+1))
                # for j in range(0,soup2num):
                #     hashtags = soup2[j].split('#')[1].split('</a>')[0]
                #     # print("hashtags: ", hashtags)
                #     csvtext.append(hashtags)
                #
                # value = ",".join(csvtext)
                # print("value: " , value)
                # sheet.append([value])

                # time.sleep(2)

    except Exception as error:
        print("오류발생: " + str(error))
        print(traceback.format_exc())


    # try:
    #     for i in range(0,reallinknum):
    #         csvtext.append([])
    #         req = Request('https://www.instagram.com/p'+reallink[i],headers={'User-Agent': 'Mozilla/5.0'})
    #
    #         webpage = urlopen(req).read()
    #         soup = BeautifulSoup(webpage,"lxml",from_encoding='utf-8')
    #         soup1 = soup.find("meta",attrs={"property":"og:description"})
    #
    #         csvtext[i].append(str(i+1))
    #         reallink1 = soup1['content']
    #         reallink1 = reallink1[reallink1.find("@")+1:reallink1.find(")")]
    #         reallink1 = reallink1[:20]
    #         if reallink1 == '':
    #             reallink1 = 'Null'
    #
    #         csvtext[i].append(reallink1)
    #         # sheet.append([reallink1])
    #
    #         for reallink2 in soup.find_all("meta",attrs={"property":"instapp:hashtags"}):
    #             reallink2 = reallink2['content']
    #             csvtext[i].append(reallink2)
    #             # sheet.append([reallink2])
    #
    #
    #         print(str(i+1)+"개의 데이터 받아오는 중.")
    #         # csvtext = [['erna_limdaugh', 'photozone', '청년몰', 'arirangtv', '전주여행', 'handmadestores', 'rangmate', 'touristspot']]
    #         value = ",".join(csvtext[i])
    #         print("value: " , value)
    #         sheet.append([value])
    #
    #         # present_date = str(datetime.utcnow() + timedelta(hours=9))[:10] #파일명에 날짜구분하기 위한 시간
    #
    #         # nowDatetime = now.strftime('%Y-%m-%d-%H-%M-%S')
    #         # data.to_csv(name + nowDatetime + ".csv", encoding='utf-8')
    #
    # except Exception as error:
    #     print("오류발생: " + str(error))

    # except:
    #     print("오류발생"+str(i+1)+"개의 데이터를 저장합니다.")

    # data = pd.DataFrame(csvtext)

    nowDatetime = now.strftime('%Y-%m-%d-%H-%M-%S')
    wb.save(name + nowDatetime + ".xlsx")
    print("All 저장성공")
    # driver.quit()
